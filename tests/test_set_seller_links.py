import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from megamarket.domain import CardToPars, Stock
from megamarket.storage.report import ExcelReport
from set_seller_links import read_path, set_seller_links


class _FakeBrowser:
    async def new_page(self):
        return object()


class _InterruptedCardParser:
    def __init__(self, page, cards) -> None:
        self.cards = cards

    async def parse_all(self):
        self.cards[0].seller_link = "https://megamarket.ru/shop/partial/"
        raise RuntimeError("Обход карточек прерван")


class SetSellerLinksTests(unittest.IsolatedAsyncioTestCase):
    def test_reads_path_from_input_when_argument_is_omitted(self):
        with (
            patch("sys.argv", ["set_seller_links.py"]),
            patch("builtins.input", return_value="output/report.xlsx"),
        ):
            self.assertEqual(read_path(), Path("output/report.xlsx"))

    async def test_saves_partial_report_when_parsing_fails(self):
        cards = [
            CardToPars(
                title="Товар",
                price="100 ₽",
                seller="Продавец",
                card_link="https://megamarket.ru/product",
                stock=Stock.IN_STOCK,
            )
        ]

        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "report.xlsx"
            ExcelReport(cards, model=CardToPars).save(source)

            with (
                patch(
                    "set_seller_links.connect_browser",
                    return_value=_FakeBrowser(),
                ),
                patch(
                    "set_seller_links.MegamarketParseCard",
                    _InterruptedCardParser,
                ),
            ):
                with self.assertRaisesRegex(RuntimeError, "прерван"):
                    await set_seller_links(source)

            output = Path(directory) / "report_with_seller_links.xlsx"
            self.assertTrue(output.is_file())
            sheet = load_workbook(output)[ExcelReport.SHEET_TITLE]
            headers = {cell.value: cell.column for cell in sheet[1]}
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(
                sheet.cell(
                    row=2,
                    column=headers["Ссылка на продавца"],
                ).value,
                "https://megamarket.ru/shop/partial/",
            )

if __name__ == "__main__":
    unittest.main()
