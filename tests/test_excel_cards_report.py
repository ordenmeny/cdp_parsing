import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from domain import CardToPars, Stock
from report import ExcelCardsReport, ExcelReport


class ExcelCardsReportTests(unittest.TestCase):
    @staticmethod
    def _cards() -> list[CardToPars]:
        return [
            CardToPars(
                title="Товар 1",
                price="100 ₽",
                seller="Продавец 1",
                card_link="https://megamarket.ru/product-1",
                stock=Stock.IN_STOCK,
            ),
            CardToPars(
                title="Товар 2",
                price="200 ₽",
                seller="Продавец 2",
                card_link="https://megamarket.ru/product-2",
                stock=Stock.IN_STOCK,
                seller_link="https://megamarket.ru/shop/already-filled/",
            ),
            CardToPars(
                title="Товар 3",
                price="",
                seller="Продавец 3",
                card_link="https://megamarket.ru/product-3",
                stock=Stock.OUT_OF_STOCK,
            ),
        ]

    def test_preserves_rows_and_existing_links(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "report.xlsx"
            ExcelReport(
                self._cards(),
                model=CardToPars,
                query="test",
            ).save(source)

            report = ExcelCardsReport(source)
            report.cards[0].seller_link = "https://megamarket.ru/shop/new-link/"
            report.cards[1].seller_link = "https://megamarket.ru/shop/changed/"
            output = report.save()

            self.assertEqual(
                output,
                Path(directory) / "report_with_seller_links.xlsx",
            )
            workbook = load_workbook(output)
            sheet = workbook[ExcelReport.SHEET_TITLE]
            headers = {
                cell.value: cell.column
                for cell in sheet[1]
                if cell.value is not None
            }
            seller_column = headers["Ссылка на продавца"]

            self.assertEqual(sheet.max_row, 4)
            self.assertEqual(
                sheet.cell(row=2, column=seller_column).value,
                "https://megamarket.ru/shop/new-link/",
            )
            self.assertEqual(
                sheet.cell(row=3, column=seller_column).value,
                "https://megamarket.ru/shop/already-filled/",
            )
            self.assertIsNone(sheet.cell(row=4, column=seller_column).value)


if __name__ == "__main__":
    unittest.main()
