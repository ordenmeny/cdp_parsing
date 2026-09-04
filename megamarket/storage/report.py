import re
from copy import copy
from collections.abc import Sequence
from datetime import datetime
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from megamarket.config import settings
from megamarket.domain import CardToPars


class ExcelReport:
    SHEET_TITLE = "Карточки"
    MAX_WIDTH = 90

    def __init__(
            self,
            rows: Sequence[BaseModel],
            *,
            model: type[BaseModel],
            query: str = "",
    ) -> None:
        self.rows = rows
        self.model = model
        self.query = query

    @staticmethod
    def _text(value) -> str:
        """В отчёте всё строками: сортировать и считать в нём нечего."""
        if isinstance(value, bool):
            return "да" if value else "нет"
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def _safe_name(value: str, fallback: str) -> str:
        return re.sub(r"[^\w-]+", "_", value).strip("_") or fallback

    @classmethod
    def _site_name(cls) -> str:
        host = urlsplit(settings.base_url).hostname or "site"
        host = host.removeprefix("www.")
        return cls._safe_name(host.split(".")[0], "site")

    def _default_path(self) -> Path:
        # Микросекунды гарантируют отдельное имя даже при быстром перезапуске.
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")
        query = self._safe_name(self.query, "query")
        directory = settings.report_dir / f"{self._site_name()}-{query}"
        return directory / f"{stamp}-{query}.xlsx"

    def save(self, path: str | Path | None = None) -> Path:
        target = Path(path) if path else self._default_path()
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.SHEET_TITLE

        # колонки - поля модели в порядке объявления
        columns = self.model.model_fields

        for number, (name, column) in enumerate(columns.items(), start=1):
            cell = sheet.cell(row=1, column=number)
            cell.value = column.title or name
            cell.font = Font(bold=True)

        for row, item in enumerate(self.rows, start=2):
            for number, name in enumerate(columns, start=1):
                value = self._text(getattr(item, name))
                sheet.cell(row=row, column=number, value=value)

        for number in range(1, len(columns) + 1):
            letter = get_column_letter(number)
            longest = max(len(str(cell.value or "")) for cell in sheet[letter])
            sheet.column_dimensions[letter].width = min(longest + 2, self.MAX_WIDTH)

        sheet.freeze_panes = "A2"
        last_column = get_column_letter(len(columns))
        sheet.auto_filter.ref = f"A1:{last_column}{len(self.rows) + 1}"

        workbook.save(target)
        print(f"Отчёт сохранён: {target.resolve()} (строк: {len(self.rows)})")
        return target


def join_excel_reports(directory: str | Path) -> Path:
    """Объединить строки всех отчётов ``.xlsx`` в указанной папке."""
    source_dir = Path(directory).expanduser().resolve()
    if not source_dir.is_dir():
        raise NotADirectoryError(f"Папка с отчётами не найдена: {source_dir}")

    files = sorted(
        path
        for path in source_dir.glob("*.xlsx")
        if path.is_file()
        and not path.name.startswith(("all-", "~$"))
    )
    if not files:
        raise FileNotFoundError(f"В папке нет файлов .xlsx: {source_dir}")

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S-%f")
    site = ExcelReport._site_name()
    directory_prefix = f"{site}-"
    query = (
        source_dir.name[len(directory_prefix):]
        if source_dir.name.startswith(directory_prefix)
        else source_dir.name
    )
    output_path = source_dir / f"all-{site}{stamp}-{query}.xlsx"
    result = Workbook()
    result_sheet = result.active
    result_sheet.title = ExcelReport.SHEET_TITLE
    expected_headers: list[str] | None = None
    output_row = 1

    for source_path in files:
        workbook = load_workbook(source_path, read_only=True, data_only=False)
        try:
            sheet = (
                workbook[ExcelReport.SHEET_TITLE]
                if ExcelReport.SHEET_TITLE in workbook.sheetnames
                else workbook.active
            )
            rows = sheet.iter_rows(values_only=True)
            headers = [str(value or "") for value in next(rows, ())]
            if not headers:
                continue
            if expected_headers is None:
                expected_headers = headers
                result_sheet.append(headers)
                output_row += 1
            elif headers != expected_headers:
                raise ValueError(
                    f"Структура колонок в файле {source_path.name} отличается."
                )

            for values in rows:
                if any(value not in (None, "") for value in values):
                    result_sheet.append(values)
                    output_row += 1
        finally:
            workbook.close()

    if expected_headers is None:
        raise ValueError("В найденных файлах отсутствуют заголовки таблицы.")

    for number in range(1, len(expected_headers) + 1):
        cell = result_sheet.cell(row=1, column=number)
        cell.font = Font(bold=True)
        letter = get_column_letter(number)
        longest = max(len(str(cell.value or "")) for cell in result_sheet[letter])
        result_sheet.column_dimensions[letter].width = min(
            longest + 2,
            ExcelReport.MAX_WIDTH,
        )

    result_sheet.freeze_panes = "A2"
    last_column = get_column_letter(len(expected_headers))
    result_sheet.auto_filter.ref = f"A1:{last_column}{output_row - 1}"
    result.save(output_path)
    print(
        f"Общий отчёт сохранён: {output_path} "
        f"(файлов: {len(files)}, строк: {output_row - 2})"
    )
    return output_path


class ExcelCardsReport:
    """Существующий отчёт с привязкой объектов ``CardToPars`` к строкам."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path).expanduser().resolve()
        if not self.path.is_file():
            raise FileNotFoundError(f"Файл отчёта не найден: {self.path}")
        if self.path.suffix.casefold() != ".xlsx":
            raise ValueError("Ожидается файл с расширением .xlsx")

        self.workbook = load_workbook(self.path)
        self.sheet = self._get_cards_sheet()
        self._columns = self._read_columns()
        self._seller_link_column = self._ensure_seller_link_column()
        self._rows: list[tuple[int, CardToPars, str]] = []
        self.cards = self._read_cards()

    def _get_cards_sheet(self):
        if ExcelReport.SHEET_TITLE in self.workbook.sheetnames:
            return self.workbook[ExcelReport.SHEET_TITLE]
        return self.workbook.active

    def _read_columns(self) -> dict[str, int]:
        columns: dict[str, int] = {}
        for cell in self.sheet[1]:
            if cell.value is not None:
                columns[str(cell.value).strip()] = cell.column
        return columns

    @staticmethod
    def _field_title(name: str) -> str:
        field = CardToPars.model_fields[name]
        return field.title or name

    def _ensure_seller_link_column(self) -> int:
        title = self._field_title("seller_link")
        existing = self._columns.get(title)
        if existing is not None:
            return existing

        column = self.sheet.max_column + 1
        cell = self.sheet.cell(row=1, column=column, value=title)
        if column > 1:
            previous = self.sheet.cell(row=1, column=column - 1)
            cell._style = copy(previous._style)
        else:
            cell.font = Font(bold=True)
        self._columns[title] = column
        return column

    def _required_columns(self) -> dict[str, int]:
        """Найти доступные колонки и проверить обязательные поля модели.

        Поля со значением по умолчанию могут отсутствовать в старых отчётах.
        Это позволяет дополнять ссылками продавцов файлы, созданные до
        появления колонки со ссылкой на изображение.
        """
        result: dict[str, int] = {}
        missing: list[str] = []
        for name, field in CardToPars.model_fields.items():
            if name == "seller_link":
                continue
            title = self._field_title(name)
            column = self._columns.get(title)
            if column is None:
                if field.is_required():
                    missing.append(title)
            else:
                result[name] = column
        if missing:
            raise ValueError(
                "В отчёте отсутствуют обязательные колонки: " + ", ".join(missing)
            )
        return result

    def _read_cards(self) -> list[CardToPars]:
        required = self._required_columns()
        cards: list[CardToPars] = []
        for row_number in range(2, self.sheet.max_row + 1):
            values = {
                name: self.sheet.cell(row=row_number, column=column).value
                for name, column in required.items()
            }
            if not any(value not in (None, "") for value in values.values()):
                continue

            data = {
                name: value
                for name, value in values.items()
                if value not in (None, "")
            }
            data["seller_link"] = (
                    self.sheet.cell(
                        row=row_number,
                        column=self._seller_link_column,
                    ).value
                    or ""
            )

            try:
                card = CardToPars.model_validate(data)
            except Exception as error:
                raise ValueError(
                    f"Не удалось прочитать строку {row_number}: {error}"
                ) from error

            original_link = card.seller_link
            self._rows.append((row_number, card, original_link))
            cards.append(card)
        return cards

    @property
    def output_path(self) -> Path:
        return self.path.with_name(f"{self.path.stem}_with_seller_links.xlsx")

    def save(
            self,
            path: str | Path | None = None,
            *,
            replace_seller_links: bool = False,
    ) -> Path:
        """Сохранить отчёт, при необходимости заменив все ссылки продавцов."""
        target = Path(path).expanduser().resolve() if path else self.output_path
        target.parent.mkdir(parents=True, exist_ok=True)

        for row_number, card, original_link in self._rows:
            if not replace_seller_links and (original_link or not card.seller_link):
                continue
            self.sheet.cell(
                row=row_number,
                column=self._seller_link_column,
                value=card.seller_link or None,
            )

        self.workbook.save(target)
        print(f"Отчёт со ссылками сохранён: {target} (строк: {len(self.cards)})")
        return target

    def close(self) -> None:
        self.workbook.close()
