import re
from copy import copy
from collections.abc import Sequence
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from config import settings
from domain import CardToPars


class ExcelReport:
    SHEET_TITLE = "Карточки"
    MAX_WIDTH = 90

    def __init__(
            self,
            rows: Sequence[BaseModel],
            *,
            model: type[BaseModel],
            query: str = "",
    ) -> None:
        self.rows = rows
        self.model = model
        self.query = query

    @staticmethod
    def _text(value) -> str:
        """В отчёте всё строками: сортировать и считать в нём нечего."""
        if isinstance(value, bool):
            return "да" if value else "нет"
        if value is None:
            return ""
        return str(value)

    def _default_path(self) -> Path:
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
        query = re.sub(r"[^\w-]+", "_", self.query).strip("_")
        name = f"megamarket_{query}_{stamp}" if query else f"megamarket_{stamp}"
        return settings.report_dir / f"{name}.xlsx"

    def save(self, path: str | Path | None = None) -> Path:
        target = Path(path) if path else self._default_path()
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self.SHEET_TITLE

        # колонки - поля модели в порядке объявления
        columns = self.model.model_fields

        for number, (name, column) in enumerate(columns.items(), start=1):
            cell = sheet.cell(row=1, column=number)
            cell.value = column.title or name
            cell.font = Font(bold=True)

        for row, item in enumerate(self.rows, start=2):
            for number, name in enumerate(columns, start=1):
                value = self._text(getattr(item, name))
                sheet.cell(row=row, column=number, value=value)

        for number in range(1, len(columns) + 1):
            letter = get_column_letter(number)
            longest = max(len(str(cell.value or "")) for cell in sheet[letter])
            sheet.column_dimensions[letter].width = min(longest + 2, self.MAX_WIDTH)

        sheet.freeze_panes = "A2"
        last_column = get_column_letter(len(columns))
        sheet.auto_filter.ref = f"A1:{last_column}{len(self.rows) + 1}"

        workbook.save(target)
        print(f"Отчёт сохранён: {target.resolve()} (строк: {len(self.rows)})")
        return target


class ExcelCardsReport:
    """Существующий отчёт с привязкой объектов ``CardToPars`` к строкам."""

    def __init__(self, path: str | Path) -> None:
        self.path = Path(path).expanduser().resolve()
        if not self.path.is_file():
            raise FileNotFoundError(f"Файл отчёта не найден: {self.path}")
        if self.path.suffix.casefold() != ".xlsx":
            raise ValueError("Ожидается файл с расширением .xlsx")

        self.workbook = load_workbook(self.path)
        self.sheet = self._get_cards_sheet()
        self._columns = self._read_columns()
        self._seller_link_column = self._ensure_seller_link_column()
        self._rows: list[tuple[int, CardToPars, str]] = []
        self.cards = self._read_cards()

    def _get_cards_sheet(self):
        if ExcelReport.SHEET_TITLE in self.workbook.sheetnames:
            return self.workbook[ExcelReport.SHEET_TITLE]
        return self.workbook.active

    def _read_columns(self) -> dict[str, int]:
        columns: dict[str, int] = {}
        for cell in self.sheet[1]:
            if cell.value is not None:
                columns[str(cell.value).strip()] = cell.column
        return columns

    @staticmethod
    def _field_title(name: str) -> str:
        field = CardToPars.model_fields[name]
        return field.title or name

    def _ensure_seller_link_column(self) -> int:
        title = self._field_title("seller_link")
        existing = self._columns.get(title)
        if existing is not None:
            return existing

        column = self.sheet.max_column + 1
        cell = self.sheet.cell(row=1, column=column, value=title)
        if column > 1:
            previous = self.sheet.cell(row=1, column=column - 1)
            cell._style = copy(previous._style)
        else:
            cell.font = Font(bold=True)
        self._columns[title] = column
        return column

    def _required_columns(self) -> dict[str, int]:
        result: dict[str, int] = {}
        missing: list[str] = []
        for name in CardToPars.model_fields:
            if name == "seller_link":
                continue
            title = self._field_title(name)
            column = self._columns.get(title)
            if column is None:
                missing.append(title)
            else:
                result[name] = column
        if missing:
            raise ValueError(
                "В отчёте отсутствуют обязательные колонки: " + ", ".join(missing)
            )
        return result

    def _read_cards(self) -> list[CardToPars]:
        required = self._required_columns()
        cards: list[CardToPars] = []
        for row_number in range(2, self.sheet.max_row + 1):
            values = {
                name: self.sheet.cell(row=row_number, column=column).value
                for name, column in required.items()
            }
            if not any(value not in (None, "") for value in values.values()):
                continue

            data = {
                "title": values.get("title") or "",
                "price": values.get("price") or "",
                "seller": values.get("seller") or "",
                "card_link": values.get("card_link") or "",
                "seller_link": (
                    self.sheet.cell(
                        row=row_number,
                        column=self._seller_link_column,
                    ).value
                    or ""
                ),
            }
            if values.get("stock") not in (None, ""):
                data["stock"] = values["stock"]

            try:
                card = CardToPars.model_validate(data)
            except Exception as error:
                raise ValueError(
                    f"Не удалось прочитать строку {row_number}: {error}"
                ) from error

            original_link = card.seller_link
            self._rows.append((row_number, card, original_link))
            cards.append(card)
        return cards

    @property
    def output_path(self) -> Path:
        return self.path.with_name(f"{self.path.stem}_with_seller_links.xlsx")

    def save(self, path: str | Path | None = None) -> Path:
        """Записать только новые ссылки, не меняя уже заполненные ячейки."""
        target = Path(path).expanduser().resolve() if path else self.output_path
        target.parent.mkdir(parents=True, exist_ok=True)

        for row_number, card, original_link in self._rows:
            if original_link or not card.seller_link:
                continue
            self.sheet.cell(
                row=row_number,
                column=self._seller_link_column,
                value=card.seller_link,
            )

        self.workbook.save(target)
        print(f"Отчёт со ссылками сохранён: {target} (строк: {len(self.cards)})")
        return target
